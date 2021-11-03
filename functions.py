"""
프로그램 관련 함수 모음집
"""
import pandas as pd  # 데이터프레임 생성용 패키지
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
data_path = "./data/"


def tf_exist_all_files():
    """
    파일 존재 여부 확인 함수
    :return:
    0(모든 파일 존재) 또는 1(대학원생 교과목 정보 파일만 존재) 또는
    2(현재 수강 교과목 정보 파일만 존재) 또는 3(필수 파일만 존재) 또는
    4(필수 파일이 존재하지 않음)
    """
    file_list = os.listdir(data_path)

    # 학부생 교과목 정보 파일(course_information_undergraduate.xls): 필수
    # 대학원생 교과목 정보 파일(course_information_graduate.xls): 필수
    # 교양 과목 정보 파일(elective_course_list.xlsx): 필수
    # 성적 정보 파일(grade_report.xls): 필수
    # 현재 수강 교과목 정보 파일(present_course_registration.xls): 선택

    tf_exist_file_e1 = "course_information_undergraduate.xls" in file_list  # 학부생 교과목 정보 파일 존재 여부
    tf_exist_file_e2 = "course_information_graduate.xls" in file_list  # 대학원생 교과목 정보 파일 존재 여부
    tf_exist_file_e3 = "elective_course_list.xlsx" in file_list  # 교양 과목 정보 파일 존재 여부
    tf_exist_file_e4 = "grade_report.xls" in file_list  # 성적 정보 파일 존재 여부
    tf_exist_file_e5 = "template.xlsm" in file_list  # 템플릿 파일 존재 여부
    tf_exist_file_e6 = "code_explain.xlsx" in file_list  # 엑셀 내 표 설명 파일 존재 여부
    tf_exist_file_s1 = "present_course_registration.xls" in file_list  # 현재 수강 교과목 정보 파일

    # 필수 파일 존재 여부 확인
    print("학부생 교과목 정보 파일:", end="\t")
    if tf_exist_file_e1:
        print("YES")
    else:
        print("NO")
        print("학부생 교과목 정보 파일(course_information_undergraduate.xls)이 존재하지 않습니다.")
        return 2

    print("대학원생 교과목 정보 파일:", end="\t")
    if tf_exist_file_e2:
        print("YES")
    else:
        print("NO")
        print("대학원생 교과목 정보 파일(course_information_graduate.xls)이 존재하지 않습니다.")
        return 2

    print("교양 과목 정보 파일:", end="\t\t")
    if tf_exist_file_e3:
        print("YES")
    else:
        print("NO")
        print("교양 과목 정보 파일(elective_course_list.xlsx)이 존재하지 않습니다.")
        return 2

    print("성적 정보 파일:", end="\t\t\t")
    if tf_exist_file_e4:
        print("YES")
    else:
        print("NO")
        print("성적 정보 파일(grade_report.xls)이 존재하지 않습니다.")
        return 2

    print("템플릿 파일:", end="\t\t\t")
    if tf_exist_file_e5:
        print("YES")
    else:
        print("NO")
        print("템플릿 파일(template.xlsm)이 존재하지 않습니다.")
        return 2

    print("테이블 설명 파일:", end="\t\t")
    if tf_exist_file_e6:
        print("YES")
    else:
        print("NO")
        print("템플릿 파일(code_explain.xlsx)이 존재하지 않습니다.")
        return 2

    # 선택 파일 존재 여부 확인
    print("현재 수강 교과목 정보 파일:", end="\t")
    if tf_exist_file_s1:
        print("YES")
    else:
        print("NO")
        return 1

    return 0


def summarize_course(ex_num):
    """
    ZEUS 상에서의 대학 및 대학원 개설강좌정보 통합 및 교과목 정리 함수
    :return:
    2016년부터 현재까지 개설된 과목 정보 dataframe
    """
    # 1. 개설강좌정보 통합
    tf_exist_graduate = ex_num == 0 or ex_num == 1  # 대학원 개설강좌정보 파일 존재 유무

    course_undergraduate = pd.read_excel(data_path + "course_information_undergraduate.xls")
    course = course_undergraduate
    # 대학원 개설강좌정보 파일이 존재할 경우에만 추가하여 합치기
    if tf_exist_graduate:
        course_graduate = pd.read_excel(data_path + "course_information_graduate.xls")
        course = pd.concat([course_undergraduate, course_graduate])

    # 2. 개설강좌정보 중 필요한 컬럼만 추출 및 컬럼 내 데이터 추출
    course = course[['년도', '학기', '교과목-분반', '교과목명', '강/실/학']].copy()

    course['전공분야코드'] = course['교과목-분반'].str[0:2]
    course['일련번호'] = course['교과목-분반'].str[2:6]
    course['학점'] = course['강/실/학'].str[-1].astype(int)
    course["교과목명"] = course["교과목명"].str.strip()

    # 3. 필요한 컬럼만 수집 및 중복되는 데이터 제거
    course = course[['년도', '학기', '전공분야코드', '일련번호', '학점', '교과목명']].drop_duplicates()

    # 4. {최초개설년도 - 교과목 정보} 형태로 묶어서 요약
    course = course.groupby(['전공분야코드', '일련번호', '학점', '교과목명'], as_index=False)[['년도', '학기']].min()
    course = course.rename(columns={'년도': '최초개설년도', '학기': '최초개설학기'})
    course = course[['전공분야코드', '일련번호', '교과목명', '최초개설년도', '최초개설학기', '학점']]

    # print(course.sort_values(by=['최초개설년도', '최초개설학기']))
    return course


def summarize_student_information(ex_num):
    """
    성적표 및 현재수강정보에서 학번 및 수강한 과목 정리
    :return:
    1. 학번 8자리
    2. 현재까지 수강 또는 이수한 과목 관련 dataframe
    """
    pd.set_option('display.max_columns', 100)  # 테이블 조회 시 테스트용 명령

    # 1. 성적표 데이터 요약
    previous = pd.read_excel(data_path + "grade_report.xls")
    # 1-1. 학번 추출
    student_number = int(previous.iloc[0][0].strip()[-8:])

    # 1-2. 의미없는 열, 행 제거
    previous = previous.drop([previous.columns[0], previous.columns[2]], axis=1)  # 열 제거
    previous = previous.drop(previous.index[:3], axis=0)  # 불필요한 부분 제거 1
    previous = previous.drop(previous.index[-3:], axis=0)  # 불필요한 부분 제거 2

    # 1-3. 인덱스 리셋 및 컬럼명 변경
    previous = previous.reset_index(drop=True)
    previous.columns = ['과목코드', '과목명', '학점', '평점']

    # 1-4. 수강학기 열 추가
    previous['수강연도'] = ""
    previous['수강학기'] = ""
    # 초기 수강 연도 및 학기 설정(AP 수강)
    year = "AP"
    semester = "AP"
    # 각 교과목의 수강 연도 및 학기 추가
    for row in range(len(previous)):
        # 연도 또는 학기가 바뀔 시 갱신
        if previous['과목명'].iloc[row].strip()[0] == "<" and previous['과목명'].iloc[row].strip()[-1] == ">":
            year_semester = previous['과목명'].iloc[row].strip()[1:-1].split("/")
            year = int(year_semester[0])  # 출력값: 2019, 2020, 2021 등
            semester = year_semester[1]  # 출력값: (한글)1학기, 여름학기,... / (영문)Spring Semester,...
        previous['수강연도'].iloc[row] = year
        previous['수강학기'].iloc[row] = semester

    # 1-5. 필요없는 행 제거
    previous = previous.dropna(subset=['과목코드'])
    previous = previous.reset_index(drop=True)

    # 1-6. 컬럼 내 데이터 추출 및 정리
    previous['전공분야코드'] = previous['과목코드'].str[0:2]
    previous['일련번호'] = previous['과목코드'].str[2:6]
    previous['학점'] = previous['학점'].astype(int)
    previous["과목명"] = previous["과목명"].str.strip()
    previous = previous[['수강연도', '수강학기', '전공분야코드', '일련번호', '과목명', '학점', '평점']]

    # 학점환산치 열 추가
    df_grade_value = pd.DataFrame({"평점": ["A+", "A0", "B+", "B0", "C+", "C0", "D+", "D0", "F", "S", "U"],
                                   "환산치": [4.5, 4.0, 3.5, 3.0, 2.5, 2.0, 1.5, 1.0, 0, "-", "-"]})
    previous = pd.merge(previous, df_grade_value, how='inner', on=["평점"])
    # 데이터를 다음과 같이 정렬
    previous = previous.sort_values(by=['수강연도', "수강학기"], axis=0)
    
    # 1-7. 정리한 성적표 데이터를 수강 기등록 과목 데이터에 넣기
    course_registration = previous

    tf_exist_present = ex_num == 0  # 현재수강과목 파일 존재 유무
    if tf_exist_present:
        # 2. 현재수강과목 데이터 요약
        present = pd.read_excel(data_path + "present_course_registration.xls")
        present = present.drop(present.index[0], axis=0)  # 맨 처음 행(header data) 제거

        # 2-1. 필요한 열만 추출, 컬럼명 변경 및 필요없는 행 제거
        present_retake = present.iloc[:, [12, 13, 14]]  # 재수강과목의 이전 수강 과목
        present_retake.columns = ['과목코드', '과목명', '드랍여부']
        present_retake = present_retake.dropna(axis=0, how='all')  # 행 전체가 결측치인 행 제거

        present = present.iloc[:, [1, 2, 7, 14]]  # 현재 수강 과목
        present.columns = ['과목코드-분반', '과목명', '강/실/학', '드랍여부']
        present = present.dropna(axis=0, how='all')  # 행 전체가 결측치인 행 제거

        # 2-2. 드랍한 교과목 제거('드랍여부' 항목이 비어있는 행만 추출)
        present = present[present['드랍여부'].isna()]
        present_retake = present_retake[present_retake['드랍여부'].isna()]

        # 2-3. 컬럼 내 데이터 추출 및 정리
        if len(present) > 0:
            present['전공분야코드'] = present['과목코드-분반'].str[0:2]
            present['일련번호'] = present['과목코드-분반'].str[2:6]
            present['학점'] = present["강/실/학"].str[-1].astype(int)
            present["과목명"] = present["과목명"].str.strip()
            present = present[['전공분야코드', '일련번호', '과목명', "학점"]]
        else:
            present['전공분야코드'] = ""
            present['일련번호'] = ""
            present['과목명'] = ""
            present['학점'] = ""
            present = present[['전공분야코드', '일련번호', '과목명', "학점"]]

        if len(present_retake) > 0:
            present_retake['전공분야코드'] = present_retake['과목코드'].str[0:2]
            present_retake['일련번호'] = present_retake['과목코드'].str[2:6]
            present_retake = present_retake[['전공분야코드', '일련번호', '과목명']]
        else:
            present_retake['전공분야코드'] = ""
            present_retake['일련번호'] = ""
            present_retake['과목명'] = ""
            present_retake = present_retake[['전공분야코드', '일련번호', '과목명']]
        
        # 2-4. 현재수강화목의 평점 및 환산치 열 추가(값은 하이픈으로 지정)
        present["평점"] = "-"
        present["환산치"] = "-"

        # 3. 성적표 데이터(previous)와 현재 수강 과목(present) 통합
        if len(present) > 0:
            course_registration = pd.concat([previous, present], ignore_index=True)

        # 4. 현재 재수강하는 과목 데이터 제거(과목코드 기반)
        for row_num in range(len(present_retake)):
            is_same_major_code = course_registration['전공분야코드'] == present_retake.iloc[row_num]['전공분야코드']
            is_same_num_code = course_registration['일련번호'] == present_retake.iloc[row_num]['일련번호']

            # 재수강하는 과목이 여러 번 이수한 과목(예체능, 콜로퀴움 등)일 경우
            # 재수강 과목 중 처음 C0 이하 또는 U로 이수한 과목 하나를 삭제하도록 조정
            tf_list = ~(is_same_major_code & is_same_num_code)
            found_removal_obj = False  # 삭제 대상 발견 tf값(발견 전: false, 발견 후: true)
            for tf_num in range(len(tf_list.tolist())):
                # 제거 대상 과목 발견 이후의 모든 과목을 그대로 보존
                if found_removal_obj:
                    tf_list[tf_num] = True
                    continue
                # 수강 과목 목론에서 제거 대상 과목 발견 시 found_removal_obj 값을 true 값으로 바꿈
                # type(tf_list[tf_num]) -> numpy.bool(bool 타입이 아니므로 not 대신 ~ 붙임)
                if ~tf_list[tf_num] and course_registration.iloc[tf_num]['평점'] in ['C0', 'D+', 'D0', 'F', 'U']:
                    found_removal_obj = True

            # 제거 대상 과목을 발견하지 못했을 시(평점이 재이수 기준 이상인 경우)
            # print 명령어로 알리고 삭제 대상 과목을 복원
            if not found_removal_obj:
                not_found_name = present_retake.iloc[row_num]['과목명']
                not_found_code = present_retake.iloc[row_num]['전공분야코드'] + present_retake.iloc[row_num]['일련번호']
                print('ALERT:  드롭 과목 중 과목명이 ' + not_found_name + '(과목 코드: ' + not_found_code + ')인 과목을 제거하지 못했습니다.')
                for i in range(len(tf_list.tolist())):
                    tf_list[i] = True
            course_registration = course_registration[tf_list]

    return student_number, course_registration


def summarize_elective_course():
    """
    elective_course_list.xlsx 내 파일 데이터 요약 함수
    :return: 
    교양 과목 관련 요약 dataframe
    """
    # 1. 분류별 데이터 추출
    hus_list = pd.read_excel(data_path + "elective_course_list.xlsx", sheet_name="hus")
    hus_list = hus_list[['교과목', '학점', '교과목명']].drop_duplicates()
    hus_list = hus_list.reset_index(drop=True)
    hus_list['분류'] = 'hus'

    ppe_list = pd.read_excel(data_path + "elective_course_list.xlsx", sheet_name="ppe")
    ppe_list = ppe_list[['교과목', '학점', '교과목명']].drop_duplicates()
    ppe_list = ppe_list.reset_index(drop=True)
    ppe_list['분류'] = 'ppe'

    gsc_list = pd.read_excel(data_path + "elective_course_list.xlsx", sheet_name="gsc")
    gsc_list = gsc_list[['교과목', '학점', '교과목명']].drop_duplicates()
    gsc_list = gsc_list.reset_index(drop=True)
    gsc_list['분류'] = 'gsc'

    # 2. gsc 관련 데이터에서 다른 분류(hus, ppe)와 중복되는 과목 삭제
    # print(len(gsc_list))
    is_gsc_hus_duplicated = pd.concat([gsc_list, hus_list]).duplicated(['교과목', '학점', '교과목명'],
                                                                       keep='last').iloc[0:len(gsc_list)]
    gsc_list = gsc_list[~is_gsc_hus_duplicated]
    # print(len(gsc_list))
    is_gsc_ppe_duplicated = pd.concat([gsc_list, ppe_list]).duplicated(['교과목', '학점', '교과목명'],
                                                                       keep='last').iloc[0:len(gsc_list)]
    gsc_list = gsc_list[~is_gsc_ppe_duplicated]
    # print(len(gsc_list))

    # 3. 모든 분류 통합
    elective_list = pd.concat([hus_list, ppe_list, gsc_list], ignore_index=True)
    elective_list['전공분야코드'] = elective_list['교과목'].str[0:2]
    elective_list['일련번호'] = elective_list['교과목'].str[2:6]
    elective_list = elective_list[['전공분야코드', '일련번호', '교과목명', '분류']]  # 학점은 학사편람 기준이므로 제외

    # print(elective_list)
    return elective_list


def excel_put_data(sheet, input_df, start_row, start_col):
    """
    데이터프레임을 엑셀에 넣기 위한 함수
    :param sheet: 데이터프레임을 입력할 시트
    :param input_df: 시트에 넣을 데이터프레임
    :param start_row: 초기 행
    :param start_col: 초기 열

    :return:
    없음(엑셀 시트에 직접 입력)
    """
    # 내용 셀의 테두리 스타일
    thin_border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))  # 좌우상하 순서

    # 정보를 각 셀에 입력
    row = start_row
    for maj_row in dataframe_to_rows(input_df, index=False, header=True):
        col = start_col
        for value in maj_row:
            sheet.cell(row=row, column=col).value = value
            # 각 셀에 테두리 추가
            if row != start_row:
                sheet.cell(row=row, column=col).border = thin_border
            # 수강횟수 셀의 경우 볼드처리
            if col == start_col+len(input_df.columns)-1:
                sheet.cell(row=row, column=col).font = Font(bold=True)
            col += 1
        row += 1


def excel_width(sheet, start_col, list_width):
    """
    엑셀 시트의 열 너비 설정 함수
    :param sheet: 너비를 설정할 시트
    :param start_col: 초기 열
    :param list_width: 설정할 열 너비 리스트

    :return:
    없음(엑셀 시트에 직접 설정)
    """
    num_list_width = len(list_width)  # 각 표의 열 개수(num...+1은 앞 빈칸을 포함한 열 개수)
    if start_col == 0:
        sheet.column_dimensions[get_column_letter(1+start_col*(num_list_width+1))].width = 1   # 빈 칸
    else:
        sheet.column_dimensions[get_column_letter(1+start_col*(num_list_width+1))].width = 3  # 빈 칸

    for i in range(num_list_width):
        sheet.column_dimensions[get_column_letter(1+start_col*(num_list_width+1) + (i+1))].width = list_width[i]


def excel_row_height(sheet):
    """
    엑셀 시트의 행 높이 설정 함수
    :param sheet: 높이를 설정할 시트

    :return:
    없음(엑셀 시트에 직접 설정)
    """
    sheet.row_dimensions[1].height = 8
    sheet.row_dimensions[2].height = 40
    sheet.row_dimensions[3].height = 8
    sheet.row_dimensions[4].height = 20
    sheet.row_dimensions[5].height = 20


def excel_design(sheet, start_col, num_columns, light_color, dark_color):
    """
    표 디자인 설정 함수
    :param sheet: 디자인 설정할 시트
    :param start_col: 초기 열
    :param num_columns: 디자인할 열 개수
    :param light_color: 밝은 색("#" 제외한 6자리 헥사코드)
    :param dark_color: 어두운 색("#" 제외한 6자리 헥사코드)

    :return:
    없음(엑셀 시트에 직접 설정)
    """
    # 셀 색상
    light = PatternFill(start_color=light_color, end_color=light_color, fill_type='solid')
    dark = PatternFill(start_color=dark_color, end_color=dark_color, fill_type='solid')

    # 셀 병합
    sheet.merge_cells(start_row=4, start_column=(num_columns+1) * start_col + 2,
                      end_row=4, end_column=(num_columns+1) * start_col + (num_columns+1))
    # 색상 설정 및 글자 서식 반영
    sheet.cell(row=4, column=(num_columns+1) * start_col + 2).fill = light
    sheet.cell(row=4, column=(num_columns+1) * start_col + 2).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
    sheet.cell(row=4, column=(num_columns+1) * start_col + 2).font = Font(bold=True)
    for i in range(2, (num_columns+1) + 1):
        sheet.cell(row=5, column=(num_columns+1) * start_col + i).fill = dark
        sheet.cell(row=5, column=(num_columns+1) * start_col + i).alignment = Alignment(horizontal='center',
                                                                                        vertical='center')
        sheet.cell(row=5, column=(num_columns+1) * start_col + i).font = Font(bold=True, color='FFFFFF')


def excel_explain_cell(sheet, str_title, str_contents, start_column, light_color):
    """
    엑셀 내 상단(헤더, ex.B2:C2 등) 설명 삽입 관련 함수
    :param sheet: 설명을 넣을 시트
    :param str_title: 제목
    :param str_contents: 내용
    :param start_column: 시작 열(행의 경우 2로 고정)
    :param light_color: 밝은 색("#" 제외한 6자리 헥사코드)

    :return:
    없음(엑셀 시트에 직접 설정)
    """
    # 셀 색상
    light = PatternFill(start_color=light_color, end_color=light_color, fill_type='solid')
    # 선 디자인
    left_border = Border(Side('thick'), Side('thin'), Side('thick'), Side('thick'))  # 좌우상하 순서
    right_border = Border(Side('thin'), Side('thick'), Side('thick'), Side('thick'))

    # 내용 반영
    sheet.cell(row=2, column=start_column).value = str_title
    sheet.cell(row=2, column=start_column+1).value = str_contents
    # 선 디자인 반영
    sheet.cell(row=2, column=start_column).border = left_border
    sheet.cell(row=2, column=start_column+1).border = right_border
    # 셀 색상 및 글꼴 굵기 반영(제목 부분)
    sheet.cell(row=2, column=start_column).fill = light
    sheet.cell(row=2, column=start_column).font = Font(bold=True)
    # 글자 서식 반영
    sheet.cell(row=2, column=start_column).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=2, column=start_column+1).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
