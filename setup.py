import pandas as pd  # 데이터프레임 생성용 패키지
import functions as func  # functions.py 파일의 함수 사용
import openpyxl

# 1. 필수 파일 존재 여부 확인
existence_number = func.tf_exist_all_files()
if existence_number == 4:
    quit(0)

# 2. 관련 변수 설정
df_course = func.summarize_course(existence_number)
student_number, df_student = func.summarize_student_information(existence_number)
df_elective = func.summarize_elective_course()
list_undergraduate_major_code = ["GS", "UC", "EC", "MA", "MC", "EV", "BS", "PS", "CH"]

# 엑셀 투입용 설명 데이터프레임
# 전공분야코드별 설명
df_major_explain = pd.read_excel("./data/code_explain.xlsx", sheet_name="major_code", keep_default_na=False)
# 혹여 추후 추가된 분야코드에서 중복되는 코드 발견 시 자동으로 삭제
df_major_explain = df_major_explain.drop_duplicates("전공분야코드")

# 교양 및 예체능 과목코드별 설명
df_elect_pna_res_explain = pd.read_excel("./data/code_explain.xlsx", sheet_name="elect_pna_res", keep_default_na=False)

# 3. 강좌개설정보에 수강한 강좌 반영
"""
코드쉐어 과목 처리 방법
1. 강좌개설정보(df_course)에서 학생 수강 과목(df_student) 중 하나와 과목코드[전공분야코드, 일련번호]가 일치하는 과목을 추출함.
2. 강좌개설정보에서 다음 조건을 만족하는 과목을 추가로 추출함.
    2-1. 1에서 추출한 과목과 [교과목명]이 같은 과목(자신 포함).
3. 2에서 만족하는 과목은 수강 횟수에 1을 더함.
4. 학생 수강 과목 내 모든 과목을 조회할 때까지 1부터 3 과정을 반복함.
"""
# 3-1. 강좌개설정보 변수(df_course)에 수강 횟수 정보 추가
df_course['수강횟수'] = 0
# 평점이 U 또는 F인 과목은 제외
df_student_not_u = df_student["평점"] != "U"
df_student_not_f = df_student["평점"] != "F"
df_student2 = df_student[df_student_not_u & df_student_not_f]

for row in range(len(df_student)):
    # 3-2. "코드쉐어 과목 처리 방법" 중 1번 항목 시행
    df_infected = df_course[df_course['전공분야코드'] == df_student2.iloc[row]['전공분야코드']]
    df_infected = df_infected[df_infected['일련번호'] == df_student2.iloc[row]['일련번호']]

    # 3-3. "코드쉐어 과목 처리 방법" 중 2번 항목 시행
    df_total_infected = pd.DataFrame()
    for inf_row in range(len(df_infected)):
        # [교과목명]이 같은 과목
        df_add_infected1 = df_course[df_course['교과목명'] == df_infected.iloc[inf_row]['교과목명']]
        df_total_infected = pd.concat([df_total_infected, df_add_infected1])
    # 중복되는 과목을 하나로 제거
    df_total_infected = df_total_infected.drop_duplicates()

    # 3-4. "코드쉐어 과목 처리 방법" 중 3번 항목 시행
    list_index_total_infected = df_total_infected.index.values
    for i in list_index_total_infected:
        df_course.loc[df_course.index == i, "수강횟수"] = df_course.loc[df_course.index == i, "수강횟수"] + 1

# 과목 수강 횟수 확인용
# print(df_course[df_course["수강횟수"] != 0].sort_values(by=['교과목명'], axis=0))


# 4. 엑셀 시트에 넣을 데이터프레임 세분화(이수과목 시트 제외)
# 4-1. 과목 전체 정보 데이터프레임 정의 및 최초개설년도 및 학기 순으로 전공분야코드 정렬
# 단, 다음 9개 코드는 순서대로 맨 앞에 정렬 - list_undergraduate_major_code(GS, UC, EC, MA, MC, EV, BS, CS, CH)
df_except_ug_major_code = df_course[~df_course["전공분야코드"].isin(list_undergraduate_major_code)]
df_except_ug_major_code = df_except_ug_major_code.groupby(["전공분야코드"], as_index=False)[["최초개설년도", "최초개설학기"]].min()
df_except_ug_major_code = df_except_ug_major_code.sort_values(by=['최초개설년도', '최초개설학기', '전공분야코드'])
# 정렬한 전공분야코드 리스트 생성
list_major_code = list_undergraduate_major_code + list(dict.fromkeys(df_except_ug_major_code["전공분야코드"].values.tolist()))


# 4-2. 교양-예체능-연구 과목 관련 데이터프레임 정의
# 4-2-1. 교양 과목 데이터프레임 정의
# df_course 변수와 합쳐 교양 과목 변수(df_course)에 수강 횟수 정보 추가
df_elective = pd.merge(df_elective, df_course, how='inner', on=["전공분야코드", "일련번호", "교과목명"])
# 같은 교과목 코드에 최신 교과목 이외 나머지 교과목을 삭제(교양 학점 계산 목적)
df_elective = df_elective.drop_duplicates(["전공분야코드", "일련번호"], keep='last')
# 컬럼명 재배열(최초개설년도, 최초개설학기 삭제)
df_elective = df_elective[["전공분야코드", "일련번호", "교과목명", "학점", "수강횟수", "분류"]]

# 4-2-2. 예체능 데이터프레임 정의
df_physical_art = df_course[df_course["전공분야코드"] == "GS"]
df_physical_art = df_physical_art[df_physical_art["일련번호"].str.startswith("01") |
                                  df_physical_art["일련번호"].str.startswith("02")]
# 컬럼명 재배열(최초개설년도, 최초개설학기 삭제)
df_physical_art = df_physical_art[["전공분야코드", "일련번호", "교과목명", "학점", "수강횟수"]]
# 같은 교과목 코드에 최신 교과목 이외 나머지 교과목을 삭제(교양 학점 계산 목적)
df_physical_art = df_physical_art.drop_duplicates(["전공분야코드", "일련번호"], keep='last')
# 예체능 과목의 "분류" 컬럼명 새로 추가
df_physical_art["분류"] = "pna"

# 4-2-3. 연구 과목 데이터프레임 정의
list_tf_res1 = df_course["일련번호"] == "9102"
list_tf_res2 = df_course["일련번호"] == "9103"
df_research = df_course[list_tf_res1 | list_tf_res2]
# 컬럼명 재배열(최초개설년도, 최초개설학기 삭제)
df_research = df_research[["전공분야코드", "일련번호", "교과목명", "학점", "수강횟수"]]
# 같은 교과목 코드에 최신 교과목 이외 나머지 교과목을 삭제(교양 학점 계산 목적)
df_research = df_research.drop_duplicates(["전공분야코드", "일련번호"], keep='last')
# 예체능 과목의 "분류" 컬럼명 새로 추가
df_research["분류"] = "res"

# 4-2-4. 교양 과목과 예체능 데이터프레임 합산
df_elect_pna_res = pd.concat([df_elective, df_physical_art, df_research])
# 교양-예체능-연구 분류코드 리스트 생성
list_elect_pna_res_code = list(dict.fromkeys(df_elect_pna_res["분류"].values.tolist()))


# 데이터프레임을 엑셀에 저장
list_format = [".xlsx", ".xlsm"]
list_use_vba = [False, True]
for i in range(len(list_format)):
    # 엑셀 파일 불러오기
    template = openpyxl.load_workbook("./data/template"+list_format[i], keep_vba=list_use_vba[i])

    # 5. 수강횟수가 반영된 강좌개설정보를 전공분야코드별로 엑셀에 저장
    # 5-1. 전공분야코드별로 템플릿 엑셀 파일에 개설강좌정보 입력
    sheet = template["전체개설과목정보"]
    # 입력 시 개설강좌정보 최좌상단 셀 위치
    start_row = 5
    start_col = 2
    # 열 개수
    num_course_columns = len(df_course.columns)  # 컬럼 개수(표 디자인용)

    for major in list_major_code:
        df_major_course = df_course[df_course["전공분야코드"] == major]  # 전공분야코드에 따라 선별한 개설강좌정보
        # 다음 조건에 따라 정렬
        df_major_course = df_major_course.sort_values(by=['일련번호', '최초개설년도', '최초개설학기'])
        # 정보를 각 셀에 입력
        func.excel_put_data(sheet=sheet, input_df=df_major_course, start_row=start_row, start_col=start_col)
        start_col += len(df_major_course.columns) + 1
        # print(df_major_course)

    # 5-2. 서식 및 디자인 설정
    # 5-2-1. 행 높이, 열 너비 설정 및 틀 고정
    # 행 높이
    func.excel_row_height(sheet)
    # 열 너비
    list_major_width = [12, 9, 35, 15, 15, 9, 9]  # 전공분야코드, 일련번호, 교과목명, 최초개설년도, 최초개설학기, 학점, 수강횟수
    for num_major in range(len(list_major_code)):
        func.excel_width(sheet=sheet, start_col=num_major, list_width=list_major_width)
    # 틀 고정
    sheet.freeze_panes = "A6"

    # 5-2-2. 강좌개설정보 부분 디자인
    for num_major in range(len(list_major_code)):
        func.excel_design(sheet=sheet, start_col=num_major, num_columns=num_course_columns,
                          light_color="B7DEE8", dark_color="31869B")

    # 5-3. 설명 부분[C2:D2] 내용 및 디자인
    cell_title = "설명"
    cell_contents = "전공분야코드 및 일련번호 중복: 있음\n교과목명 중복: 있음"
    func.excel_explain_cell(sheet=sheet, str_title=cell_title, str_contents=cell_contents,
                            start_column=3, light_color="B7DEE8")

    # 5-4. 각 전공분야코드별 설명 추가
    # 엑셀 파일에 설명 추가
    for num_major in range(len(list_major_code)):
        sheet.cell(row=4, column=(num_course_columns+1) * num_major + 2).value = \
            df_major_explain.loc[df_major_explain["전공분야코드"] == list_major_code[num_major], "설명"].values[0]

    # 6. 교양, 예체능 및 연구 과목에 수강횟수를 반영하여 엑셀에 저장
    sheet = template["교양-예체능-연구"]

    # 6-1. 분류별로 엑셀 시트에 기입
    # 입력 시 최좌상단 셀 위치
    start_row = 5
    start_col = 2
    # 열 개수
    num_elect_pna_res_columns = len(df_elect_pna_res.columns)-1  # 분류 컬럼명을 제외한컬럼 개수

    for elect in list_elect_pna_res_code:
        # 교양과목 분류에 따라 선별한 개설강좌정보
        df_elect_pna_res_course = df_elect_pna_res[df_elect_pna_res["분류"] == elect].drop(["분류"], axis=1)
        # 다음 조건에 따라 정렬
        df_elect_pna_res_course = df_elect_pna_res_course.sort_values(by=["전공분야코드", '일련번호'])
        # 컬럼명 재배열(분류 삭제)
        df_elect_pna_res_course = df_elect_pna_res_course[["전공분야코드", "일련번호", "교과목명", "학점", "수강횟수"]]
        # 정보를 각 셀에 입력
        func.excel_put_data(sheet=sheet, input_df=df_elect_pna_res_course, start_row=start_row, start_col=start_col)
        start_col += num_elect_pna_res_columns + 1
        # print(df_elective_course)

    # 6-2. 서식 및 디자인 설정
    # 6-2-1. 행 높이, 열 너비 설정 및 틀 고정
    # 행 높이
    func.excel_row_height(sheet)
    # 열 너비
    list_elect_pna_res_width = [12, 9, 35, 9, 9]  # 전공분야코드, 일련번호, 교과목명, 학점, 수강횟수
    for num_elective in range(len(list_elect_pna_res_code)):
        func.excel_width(sheet=sheet, start_col=num_elective, list_width=list_elect_pna_res_width)
    # 틀 고정
    sheet.freeze_panes = "A6"

    # 6-2-2. 강좌개설정보 부분 디자인
    for num_elect_pna_res in range(len(list_elect_pna_res_code)):
        func.excel_design(sheet=sheet, start_col=num_elect_pna_res, num_columns=num_elect_pna_res_columns,
                          light_color="E2EFDA", dark_color="548235")

    # 6-3. 설명 부분[C2:D2] 내용 및 디자인
    cell_title = "설명"
    cell_contents = "전공분야코드 및 일련번호 중복: 없음\n교과목명 중복: 없음"
    func.excel_explain_cell(sheet=sheet, str_title=cell_title, str_contents=cell_contents,
                            start_column=3, light_color="E2EFDA")
    # 6-4. 각 전공분야코드별 설명 추가
    # 엑셀 파일에 설명 추가
    for num_elect_pna_res in range(len(list_elect_pna_res_code)):
        sheet.cell(row=4, column=(num_elect_pna_res_columns+1) * num_elect_pna_res + 2).value = \
            df_elect_pna_res_explain.loc[df_elect_pna_res_explain["분류"] == list_elect_pna_res_code[num_elect_pna_res],
                                         "설명"].values[0]

    # 7. 성적 관련 정보를 엑셀에 저장
    sheet = template["수강과목요약"]

    # 7-1. 분류별로 엑셀 시트에 기입
    # 입력 시 최좌상단 셀 위치
    start_row = 5
    start_col = 2

    # 정보를 각 셀에 입력
    func.excel_put_data(sheet=sheet, input_df=df_student, start_row=start_row, start_col=start_col)

    # 7-2. 서식 및 디자인 설정
    # 7-2-1. 행 높이, 열 너비 설정 및 틀 고정
    # 행 높이
    func.excel_row_height(sheet)
    # 열 너비
    list_student_width = [15, 15, 12, 9, 45, 9, 9, 9]  # 수강연도, 수강학기, 전공분야코드, 일련번호, 과목명, 학점, 평점, 환산치
    func.excel_width(sheet=sheet, start_col=0, list_width=list_student_width)
    # 틀 고정
    sheet.freeze_panes = "A6"

    # 7-2-2. 학번 부분[B2:C2] 내용 및 디자인
    cell_title = "학번"
    func.excel_explain_cell(sheet=sheet, str_title=cell_title, str_contents=student_number,
                            start_column=2, light_color="FCE4D6")

    # 7-2-3. 강좌개설정보 부분 디자인
    # 열 개수
    num_student_columns = len(df_student.columns)  # 성적표 데이터베이스 컬럼 개수
    func.excel_design(sheet=sheet, start_col=0, num_columns=num_student_columns,
                      light_color="FCE4D6", dark_color="C65911")

    # 7-3. 엑셀 파일에 설명 추가
    sheet.cell(row=4, column=2).value = "총 수강 과목"

    # 8. 기초 및 (대학)전공과목 정보를 엑셀에 저장
    sheet = template["기초및전공과목"]
    # 8-1. 분류별로 엑셀 시트에 기입
    # 입력 시 최좌상단 셀 위치
    start_row = 5
    start_col = 2
    for under in list_undergraduate_major_code:
        # 교양과목 분류에 따라 선별한 개설강좌정보
        df_undergraduate_course = df_course[df_course["전공분야코드"] == under]
        # 다음 조건에 따라 정렬
        df_undergraduate_course = df_undergraduate_course.sort_values(by=["최초개설년도", "최초개설학기"])
        # 같은 교과목명에 최신 교과목 이외 나머지 교과목을 삭제(전공 학점 계산 목적)
        df_undergraduate_course = df_undergraduate_course.drop_duplicates(["교과목명"], keep='last')
        # GS 과목 및 UC 과목 외 대학원 과목 및 학사논문연구 교과목 제거
        if under != "GS" and under != "UC":
            df_undergraduate_course = df_undergraduate_course[df_undergraduate_course["일련번호"].str[0] <= "4"]
        # 다음 조건에 따라 정렬
        df_undergraduate_course = df_undergraduate_course.sort_values(by='일련번호')
        # 컬럼명 재배열(최초개설년도 및 학기 삭제)
        df_undergraduate_course = df_undergraduate_course[["전공분야코드", "일련번호", "교과목명", "학점", "수강횟수"]]
        # 정보를 각 셀에 입력
        func.excel_put_data(sheet=sheet, input_df=df_undergraduate_course, start_row=start_row, start_col=start_col)
        start_col += num_elect_pna_res_columns + 1
        # print(df_elective_course)

    # 8-2. 서식 및 디자인 설정
    # 8-2-1. 행 높이, 열 너비 설정 및 틀 고정
    # 행 높이
    func.excel_row_height(sheet)
    # 열 너비
    list_undergraduate_width = [12, 9, 35, 9, 9]  # 전공분야코드, 일련번호, 교과목명, 학점, 수강횟수
    num_undergraduate_columns = len(list_undergraduate_width)
    for num_undergraduate in range(len(list_undergraduate_major_code)):
        func.excel_width(sheet=sheet, start_col=num_undergraduate, list_width=list_undergraduate_width)
    # 틀 고정
    sheet.freeze_panes = "A6"

    # 8-2-2. 강좌개설정보 부분 디자인
    for num_undergraduate in range(len(list_undergraduate_major_code)):
        func.excel_design(sheet=sheet, start_col=num_undergraduate, num_columns=num_undergraduate_columns,
                          light_color="FFF2CC", dark_color="BF8F00")

    # 8-3. 설명 부분[C2:D2] 내용 및 디자인
    cell_title = "설명"
    cell_contents = "전공분야코드 및 일련번호 중복: 없음\n교과목명 중복: 없음"
    func.excel_explain_cell(sheet=sheet, str_title=cell_title, str_contents=cell_contents,
                            start_column=3, light_color="FFF2CC")

    # 8-4. 각 전공분야코드별 설명 추가
    # 엑셀 파일에 설명 추가
    for num_undergraduate in range(len(list_undergraduate_major_code)):
        sheet.cell(row=4, column=(num_undergraduate_columns+1) * num_undergraduate + 2).value = \
            df_major_explain.loc[df_major_explain["전공분야코드"] == list_undergraduate_major_code[num_undergraduate],
                                 "설명"].values[0]

    # 9. 입력한 정보를 저장
    template.save(filename="computerization_result_kor"+list_format[i])
    print("Saved as", "computerization_result_kor"+list_format[i])
